from dotenv import load_dotenv
from pathlib import Path
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import io
import math
import uuid
import bcrypt
import jwt
import logging
from datetime import datetime, timezone, timedelta, date
from typing import Optional, List, Literal, Any

from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Query
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr

from openpyxl import Workbook
from openpyxl.styles import Font as XLFont, Alignment, PatternFill
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm

# ---- DB ----
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# ---- App ----
app = FastAPI(title="SiKerja - Sistem Kehadiran Pegawai")
api = APIRouter(prefix="/api")
security = HTTPBearer(auto_error=False)

JWT_ALGO = "HS256"
JWT_SECRET = os.environ["JWT_SECRET"]

# ---- Helpers ----
def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()

def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False

def create_token(user_id: str, role: str, hours: int = 24) -> str:
    payload = {
        "sub": user_id,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=hours),
        "type": "access",
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGO)

def now_utc_iso() -> str:
    return datetime.now(timezone.utc).isoformat()

def today_str() -> str:
    return date.today().isoformat()

def haversine_m(lat1, lon1, lat2, lon2) -> float:
    R = 6371000.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlmb = math.radians(lon2 - lon1)
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlmb/2)**2
    return 2 * R * math.asin(math.sqrt(a))

def clean_doc(d: dict) -> dict:
    if not d:
        return d
    # Preserve our string UUID id; only fall back to Mongo _id if 'id' missing.
    if "_id" in d:
        mid = d.pop("_id")
        if "id" not in d:
            d["id"] = str(mid) if not isinstance(mid, str) else mid
    d.pop("password_hash", None)
    return d

async def get_current_user(request: Request, creds: Optional[HTTPAuthorizationCredentials] = Depends(security)) -> dict:
    token = None
    if creds:
        token = creds.credentials
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Tidak terautentikasi")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Sesi telah berakhir")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token tidak valid")
    user = await db.users.find_one({"id": payload["sub"]})
    if not user:
        raise HTTPException(status_code=401, detail="Pengguna tidak ditemukan")
    return clean_doc(user)

async def require_admin(user: dict = Depends(get_current_user)) -> dict:
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Akses ditolak")
    return user

# ---- Models ----
class LoginBody(BaseModel):
    email: str
    password: str

class EmployeeCreate(BaseModel):
    employee_id: str
    full_name: str
    email: EmailStr
    position: Optional[str] = ""
    department_id: Optional[str] = None
    phone: Optional[str] = ""
    profile_photo: Optional[str] = ""
    password: str = "employee123"
    status: str = "active"

class EmployeeUpdate(BaseModel):
    full_name: Optional[str] = None
    email: Optional[EmailStr] = None
    position: Optional[str] = None
    department_id: Optional[str] = None
    phone: Optional[str] = None
    profile_photo: Optional[str] = None
    status: Optional[str] = None
    password: Optional[str] = None

class DepartmentCreate(BaseModel):
    name: str
    description: Optional[str] = ""

class LocationCreate(BaseModel):
    name: str
    latitude: float
    longitude: float
    allowed_radius: int = 100
    status: str = "active"

class HolidayCreate(BaseModel):
    holiday_date: str
    name: str
    description: Optional[str] = ""

class LeaveCreate(BaseModel):
    leave_type: str
    start_date: str
    end_date: str
    reason: str
    attachment: Optional[str] = ""

class LeaveDecision(BaseModel):
    status: Literal["approved", "rejected"]

class WorkPlanCreate(BaseModel):
    work_date: str
    items: List[dict]  # [{work_plan, priority, target, notes}]

class CheckInBody(BaseModel):
    attendance_type: Literal["WFO", "WFH"]
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    accuracy: Optional[float] = None
    selfie: str  # base64
    location_id: Optional[str] = None

class CheckOutBody(BaseModel):
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    accuracy: Optional[float] = None
    selfie: str
    activity: str
    output: str
    activity_status: Literal["completed", "partial", "not_completed"] = "completed"
    notes: Optional[str] = ""

class SettingsBody(BaseModel):
    office_name: Optional[str] = None
    logo: Optional[str] = None
    main_title: Optional[str] = None
    subtitle: Optional[str] = None
    primary_color: Optional[str] = None
    secondary_color: Optional[str] = None
    ad_enabled: Optional[bool] = None
    ad_image: Optional[str] = None
    ad_title: Optional[str] = None
    ad_description: Optional[str] = None
    ad_url: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    late_threshold_minutes: Optional[int] = None
    min_work_minutes: Optional[int] = None

# ---- Auth ----
@api.post("/auth/admin/login")
async def admin_login(body: LoginBody):
    user = await db.users.find_one({"email": body.email.lower(), "role": "admin"})
    if not user or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Email atau kata sandi salah")
    token = create_token(user["id"], "admin")
    return {"token": token, "user": clean_doc(user)}

@api.post("/auth/employee/login")
async def employee_login(body: LoginBody):
    user = await db.users.find_one({"email": body.email.lower(), "role": "employee"})
    if not user or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Email atau kata sandi salah")
    if user.get("status") != "active":
        raise HTTPException(status_code=403, detail="Akun tidak aktif")
    token = create_token(user["id"], "employee")
    return {"token": token, "user": clean_doc(user)}

@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    if user.get("role") == "employee":
        emp = await db.employees.find_one({"user_id": user["id"]})
        user["employee"] = clean_doc(emp) if emp else None
    return user

@api.post("/auth/change-password")
async def change_password(body: dict, user: dict = Depends(get_current_user)):
    old = body.get("old_password")
    new = body.get("new_password")
    if not new or len(new) < 6:
        raise HTTPException(status_code=400, detail="Kata sandi baru minimal 6 karakter")
    u = await db.users.find_one({"id": user["id"]})
    if not verify_password(old or "", u["password_hash"]):
        raise HTTPException(status_code=400, detail="Kata sandi lama salah")
    await db.users.update_one({"id": user["id"]}, {"$set": {"password_hash": hash_password(new)}})
    return {"ok": True}

# ---- Departments ----
@api.get("/departments")
async def list_departments(user: dict = Depends(get_current_user)):
    items = await db.departments.find().to_list(1000)
    return [clean_doc(x) for x in items]

@api.post("/departments")
async def create_department(body: DepartmentCreate, admin: dict = Depends(require_admin)):
    doc = {"id": str(uuid.uuid4()), **body.model_dump(), "status": "active", "created_at": now_utc_iso()}
    await db.departments.insert_one(doc)
    return clean_doc(doc)

@api.put("/departments/{dept_id}")
async def update_department(dept_id: str, body: DepartmentCreate, admin: dict = Depends(require_admin)):
    await db.departments.update_one({"id": dept_id}, {"$set": body.model_dump()})
    doc = await db.departments.find_one({"id": dept_id})
    return clean_doc(doc)

@api.delete("/departments/{dept_id}")
async def delete_department(dept_id: str, admin: dict = Depends(require_admin)):
    await db.departments.delete_one({"id": dept_id})
    return {"ok": True}

# ---- Employees ----
@api.get("/employees")
async def list_employees(user: dict = Depends(get_current_user), department_id: Optional[str] = None, status: Optional[str] = None):
    q = {}
    if department_id: q["department_id"] = department_id
    if status: q["status"] = status
    items = await db.employees.find(q).to_list(2000)
    depts = {d["id"]: d.get("name", "") for d in await db.departments.find().to_list(1000)}
    out = []
    for x in items:
        c = clean_doc(x)
        c["department_name"] = depts.get(c.get("department_id"), "-")
        out.append(c)
    return out

@api.get("/employees/{emp_id}")
async def get_employee(emp_id: str, user: dict = Depends(get_current_user)):
    e = await db.employees.find_one({"id": emp_id})
    if not e: raise HTTPException(404, "Pegawai tidak ditemukan")
    return clean_doc(e)

@api.post("/employees")
async def create_employee(body: EmployeeCreate, admin: dict = Depends(require_admin)):
    if await db.users.find_one({"email": body.email.lower()}):
        raise HTTPException(400, "Email sudah terdaftar")
    if await db.employees.find_one({"employee_id": body.employee_id}):
        raise HTTPException(400, "NIP/ID Pegawai sudah terdaftar")
    user_id = str(uuid.uuid4())
    emp_id = str(uuid.uuid4())
    user_doc = {
        "id": user_id,
        "email": body.email.lower(),
        "password_hash": hash_password(body.password),
        "role": "employee",
        "status": body.status,
        "created_at": now_utc_iso(),
    }
    emp_doc = {
        "id": emp_id,
        "user_id": user_id,
        "employee_id": body.employee_id,
        "full_name": body.full_name,
        "email": body.email.lower(),
        "position": body.position,
        "department_id": body.department_id,
        "phone": body.phone,
        "profile_photo": body.profile_photo,
        "status": body.status,
        "created_at": now_utc_iso(),
    }
    await db.users.insert_one(user_doc)
    await db.employees.insert_one(emp_doc)
    return clean_doc(emp_doc)

@api.put("/employees/{emp_id}")
async def update_employee(emp_id: str, body: EmployeeUpdate, admin: dict = Depends(require_admin)):
    e = await db.employees.find_one({"id": emp_id})
    if not e: raise HTTPException(404, "Tidak ditemukan")
    updates = {k: v for k, v in body.model_dump().items() if v is not None and k != "password"}
    if updates:
        await db.employees.update_one({"id": emp_id}, {"$set": updates})
    user_updates = {}
    if body.email: user_updates["email"] = body.email.lower()
    if body.status: user_updates["status"] = body.status
    if body.password: user_updates["password_hash"] = hash_password(body.password)
    if user_updates:
        await db.users.update_one({"id": e["user_id"]}, {"$set": user_updates})
    doc = await db.employees.find_one({"id": emp_id})
    return clean_doc(doc)

@api.delete("/employees/{emp_id}")
async def delete_employee(emp_id: str, admin: dict = Depends(require_admin)):
    e = await db.employees.find_one({"id": emp_id})
    if not e: raise HTTPException(404, "Tidak ditemukan")
    await db.users.delete_one({"id": e["user_id"]})
    await db.employees.delete_one({"id": emp_id})
    return {"ok": True}

# ---- Locations ----
@api.get("/locations")
async def list_locations(user: dict = Depends(get_current_user)):
    return [clean_doc(x) for x in await db.locations.find().to_list(100)]

@api.post("/locations")
async def create_location(body: LocationCreate, admin: dict = Depends(require_admin)):
    doc = {"id": str(uuid.uuid4()), **body.model_dump(), "created_at": now_utc_iso()}
    await db.locations.insert_one(doc)
    return clean_doc(doc)

@api.put("/locations/{loc_id}")
async def update_location(loc_id: str, body: LocationCreate, admin: dict = Depends(require_admin)):
    await db.locations.update_one({"id": loc_id}, {"$set": body.model_dump()})
    return clean_doc(await db.locations.find_one({"id": loc_id}))

@api.delete("/locations/{loc_id}")
async def delete_location(loc_id: str, admin: dict = Depends(require_admin)):
    await db.locations.delete_one({"id": loc_id})
    return {"ok": True}

# ---- Holidays ----
@api.get("/holidays")
async def list_holidays(user: dict = Depends(get_current_user)):
    items = await db.holidays.find().sort("holiday_date", 1).to_list(1000)
    return [clean_doc(x) for x in items]

@api.post("/holidays")
async def create_holiday(body: HolidayCreate, admin: dict = Depends(require_admin)):
    doc = {"id": str(uuid.uuid4()), **body.model_dump(), "created_at": now_utc_iso()}
    await db.holidays.insert_one(doc)
    return clean_doc(doc)

@api.delete("/holidays/{hid}")
async def delete_holiday(hid: str, admin: dict = Depends(require_admin)):
    await db.holidays.delete_one({"id": hid})
    return {"ok": True}

# ---- Settings ----
DEFAULT_SETTINGS = {
    "id": "app",
    "office_name": "Kantor Pemerintahan",
    "logo": "",
    "main_title": "SiKerja",
    "subtitle": "Sistem Kehadiran Pegawai",
    "primary_color": "#800000",
    "secondary_color": "#f8f9fa",
    "ad_enabled": True,
    "ad_image": "https://images.pexels.com/photos/16898413/pexels-photo-16898413.jpeg",
    "ad_title": "Layanan Publik Pemerintah",
    "ad_description": "Informasi layanan publik terbaru dari pemerintah.",
    "ad_url": "",
    "start_time": "07:30",
    "end_time": "16:00",
    "late_threshold_minutes": 15,
    "min_work_minutes": 420,
}

@api.get("/settings")
async def get_settings():
    s = await db.settings.find_one({"id": "app"})
    if not s:
        await db.settings.insert_one(DEFAULT_SETTINGS.copy())
        s = DEFAULT_SETTINGS.copy()
    s.pop("_id", None)
    return s

@api.put("/settings")
async def update_settings(body: SettingsBody, admin: dict = Depends(require_admin)):
    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    await db.settings.update_one({"id": "app"}, {"$set": updates}, upsert=True)
    s = await db.settings.find_one({"id": "app"})
    s.pop("_id", None)
    return s

# ---- Work plans ----
@api.post("/work-plans")
async def create_work_plan(body: WorkPlanCreate, user: dict = Depends(get_current_user)):
    emp = await db.employees.find_one({"user_id": user["id"]})
    if not emp: raise HTTPException(400, "Data pegawai tidak ditemukan")
    # Delete existing for same date + employee, then insert new items
    await db.work_plans.delete_many({"employee_id": emp["id"], "work_date": body.work_date})
    docs = []
    for item in body.items:
        docs.append({
            "id": str(uuid.uuid4()),
            "employee_id": emp["id"],
            "work_date": body.work_date,
            "work_plan": item.get("work_plan", ""),
            "priority": item.get("priority", "normal"),
            "target": item.get("target", ""),
            "notes": item.get("notes", ""),
            "created_at": now_utc_iso(),
        })
    if docs:
        await db.work_plans.insert_many(docs)
    return {"ok": True, "count": len(docs)}

@api.get("/work-plans")
async def list_work_plans(user: dict = Depends(get_current_user), employee_id: Optional[str] = None, work_date: Optional[str] = None):
    q = {}
    if user.get("role") == "employee":
        emp = await db.employees.find_one({"user_id": user["id"]})
        q["employee_id"] = emp["id"] if emp else "-"
    elif employee_id:
        q["employee_id"] = employee_id
    if work_date:
        q["work_date"] = work_date
    items = await db.work_plans.find(q).sort("work_date", -1).to_list(500)
    return [clean_doc(x) for x in items]

# ---- Leaves ----
@api.post("/leaves")
async def create_leave(body: LeaveCreate, user: dict = Depends(get_current_user)):
    emp = await db.employees.find_one({"user_id": user["id"]})
    if not emp: raise HTTPException(400, "Data pegawai tidak ditemukan")
    doc = {
        "id": str(uuid.uuid4()),
        "employee_id": emp["id"],
        **body.model_dump(),
        "status": "pending",
        "created_at": now_utc_iso(),
    }
    await db.leaves.insert_one(doc)
    return clean_doc(doc)

@api.get("/leaves")
async def list_leaves(user: dict = Depends(get_current_user), status: Optional[str] = None, employee_id: Optional[str] = None):
    q = {}
    if user.get("role") == "employee":
        emp = await db.employees.find_one({"user_id": user["id"]})
        q["employee_id"] = emp["id"] if emp else "-"
    elif employee_id:
        q["employee_id"] = employee_id
    if status: q["status"] = status
    items = await db.leaves.find(q).sort("created_at", -1).to_list(500)
    emps = {e["id"]: e.get("full_name", "") for e in await db.employees.find().to_list(2000)}
    out = []
    for x in items:
        c = clean_doc(x)
        c["employee_name"] = emps.get(c.get("employee_id"), "-")
        out.append(c)
    return out

@api.put("/leaves/{lid}/decision")
async def decide_leave(lid: str, body: LeaveDecision, admin: dict = Depends(require_admin)):
    await db.leaves.update_one({"id": lid}, {"$set": {"status": body.status, "approved_by": admin["id"], "approved_at": now_utc_iso()}})
    return clean_doc(await db.leaves.find_one({"id": lid}))

# ---- Attendance ----
async def _get_settings_dict():
    s = await db.settings.find_one({"id": "app"}) or DEFAULT_SETTINGS.copy()
    s.pop("_id", None)
    return s

def _compute_status(check_in_time: str, settings: dict) -> str:
    try:
        t = datetime.fromisoformat(check_in_time)
        sh, sm = [int(x) for x in settings.get("start_time", "07:30").split(":")]
        threshold = timedelta(minutes=int(settings.get("late_threshold_minutes", 15)))
        official = t.replace(hour=sh, minute=sm, second=0, microsecond=0)
        return "late" if t > (official + threshold) else "on_time"
    except Exception:
        return "on_time"

@api.post("/attendance/check-in")
async def check_in(body: CheckInBody, user: dict = Depends(get_current_user)):
    if user.get("role") != "employee":
        raise HTTPException(403, "Hanya pegawai yang dapat melakukan check-in")
    emp = await db.employees.find_one({"user_id": user["id"]})
    if not emp: raise HTTPException(400, "Data pegawai tidak ditemukan")
    tdy = today_str()
    # Prevent duplicate check-in
    existing = await db.attendance.find_one({"employee_id": emp["id"], "attendance_date": tdy})
    if existing and existing.get("check_in_time"):
        raise HTTPException(400, "Anda sudah melakukan check-in hari ini")
    # Check leave/holiday
    if await db.holidays.find_one({"holiday_date": tdy}):
        raise HTTPException(400, "Hari ini adalah hari libur")
    lv = await db.leaves.find_one({"employee_id": emp["id"], "status": "approved", "start_date": {"$lte": tdy}, "end_date": {"$gte": tdy}})
    if lv:
        raise HTTPException(400, "Anda sedang cuti hari ini")

    gps_verified = False
    distance = None
    if body.attendance_type == "WFO":
        if body.latitude is None or body.longitude is None:
            raise HTTPException(400, "Lokasi GPS diperlukan untuk WFO")
        locs = await db.locations.find({"status": "active"}).to_list(50)
        if not locs:
            raise HTTPException(400, "Belum ada lokasi kantor terdaftar")
        best = None
        for loc in locs:
            d = haversine_m(body.latitude, body.longitude, loc["latitude"], loc["longitude"])
            if best is None or d < best[0]:
                best = (d, loc)
        distance, loc = best
        if distance <= loc.get("allowed_radius", 100):
            gps_verified = True
        else:
            raise HTTPException(400, f"Anda berada {int(distance)}m dari kantor {loc['name']} (radius {loc['allowed_radius']}m). Absensi tidak dapat direkam.")
    else:  # WFH
        gps_verified = body.latitude is not None and body.longitude is not None

    settings = await _get_settings_dict()
    now_iso = now_utc_iso()
    status = _compute_status(now_iso, settings)

    doc = {
        "id": str(uuid.uuid4()),
        "employee_id": emp["id"],
        "attendance_date": tdy,
        "attendance_type": body.attendance_type,
        "check_in_time": now_iso,
        "check_out_time": None,
        "check_in_latitude": body.latitude,
        "check_in_longitude": body.longitude,
        "check_in_accuracy": body.accuracy,
        "check_in_selfie": body.selfie,
        "check_in_distance": distance,
        "gps_verified": gps_verified,
        "attendance_status": status,
        "check_out_selfie": None,
        "check_out_latitude": None,
        "check_out_longitude": None,
        "activity": None,
        "output": None,
        "activity_status": None,
        "notes": None,
        "created_at": now_iso,
    }
    if existing:
        await db.attendance.update_one({"id": existing["id"]}, {"$set": doc})
    else:
        await db.attendance.insert_one(doc)
    return clean_doc(doc)

@api.post("/attendance/check-out")
async def check_out(body: CheckOutBody, user: dict = Depends(get_current_user)):
    if user.get("role") != "employee":
        raise HTTPException(403, "Hanya pegawai")
    emp = await db.employees.find_one({"user_id": user["id"]})
    if not emp: raise HTTPException(400, "Data pegawai tidak ditemukan")
    tdy = today_str()
    att = await db.attendance.find_one({"employee_id": emp["id"], "attendance_date": tdy})
    if not att or not att.get("check_in_time"):
        raise HTTPException(400, "Anda belum melakukan check-in hari ini")
    if att.get("check_out_time"):
        raise HTTPException(400, "Anda sudah melakukan check-out hari ini")

    now_iso = now_utc_iso()
    updates = {
        "check_out_time": now_iso,
        "check_out_latitude": body.latitude,
        "check_out_longitude": body.longitude,
        "check_out_accuracy": body.accuracy,
        "check_out_selfie": body.selfie,
        "activity": body.activity,
        "output": body.output,
        "activity_status": body.activity_status,
        "notes": body.notes,
    }
    await db.attendance.update_one({"id": att["id"]}, {"$set": updates})
    return clean_doc({**att, **updates})

@api.get("/attendance/today")
async def attendance_today(user: dict = Depends(get_current_user)):
    if user.get("role") != "employee":
        raise HTTPException(403, "Hanya pegawai")
    emp = await db.employees.find_one({"user_id": user["id"]})
    if not emp: return None
    att = await db.attendance.find_one({"employee_id": emp["id"], "attendance_date": today_str()})
    return clean_doc(att) if att else None

@api.get("/attendance")
async def list_attendance(
    user: dict = Depends(get_current_user),
    employee_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    department_id: Optional[str] = None,
    attendance_type: Optional[str] = None,
    status: Optional[str] = None,
):
    q: dict = {}
    if user.get("role") == "employee":
        emp = await db.employees.find_one({"user_id": user["id"]})
        q["employee_id"] = emp["id"] if emp else "-"
    else:
        if employee_id: q["employee_id"] = employee_id
        if department_id:
            emp_ids = [e["id"] for e in await db.employees.find({"department_id": department_id}).to_list(2000)]
            q["employee_id"] = {"$in": emp_ids}
    if date_from or date_to:
        q["attendance_date"] = {}
        if date_from: q["attendance_date"]["$gte"] = date_from
        if date_to: q["attendance_date"]["$lte"] = date_to
    if attendance_type: q["attendance_type"] = attendance_type
    if status: q["attendance_status"] = status
    items = await db.attendance.find(q).sort("attendance_date", -1).to_list(5000)
    emps = {e["id"]: e for e in await db.employees.find().to_list(2000)}
    depts = {d["id"]: d.get("name", "") for d in await db.departments.find().to_list(1000)}
    out = []
    for x in items:
        c = clean_doc(x)
        e = emps.get(c.get("employee_id")) or {}
        c["employee_name"] = e.get("full_name", "-")
        c["employee_code"] = e.get("employee_id", "-")
        c["department_name"] = depts.get(e.get("department_id"), "-")
        out.append(c)
    return out

@api.get("/attendance/{att_id}")
async def get_attendance(att_id: str, user: dict = Depends(get_current_user)):
    a = await db.attendance.find_one({"id": att_id})
    if not a: raise HTTPException(404, "Tidak ditemukan")
    if user.get("role") == "employee":
        emp = await db.employees.find_one({"user_id": user["id"]})
        if not emp or a["employee_id"] != emp["id"]:
            raise HTTPException(403, "Akses ditolak")
    return clean_doc(a)

# ---- Dashboard Stats ----
@api.get("/dashboard/stats")
async def dashboard_stats(admin: dict = Depends(require_admin)):
    tdy = today_str()
    total_employees = await db.employees.count_documents({"status": "active"})
    today_att = await db.attendance.find({"attendance_date": tdy}).to_list(5000)
    wfo = sum(1 for a in today_att if a.get("attendance_type") == "WFO")
    wfh = sum(1 for a in today_att if a.get("attendance_type") == "WFH")
    late = sum(1 for a in today_att if a.get("attendance_status") == "late")
    not_checked_out = sum(1 for a in today_att if a.get("check_in_time") and not a.get("check_out_time"))
    present = len(today_att)
    approved_leaves = await db.leaves.count_documents({"status": "approved", "start_date": {"$lte": tdy}, "end_date": {"$gte": tdy}})
    absent = max(0, total_employees - present - approved_leaves)

    # Trend last 7 days
    trend = []
    for i in range(6, -1, -1):
        d = (date.today() - timedelta(days=i)).isoformat()
        c = await db.attendance.count_documents({"attendance_date": d})
        trend.append({"date": d, "count": c})

    # Department comparison
    depts = await db.departments.find().to_list(100)
    dept_stats = []
    for dep in depts:
        emp_ids = [e["id"] for e in await db.employees.find({"department_id": dep["id"]}).to_list(500)]
        c = await db.attendance.count_documents({"attendance_date": tdy, "employee_id": {"$in": emp_ids}})
        dept_stats.append({"name": dep["name"], "present": c, "total": len(emp_ids)})

    return {
        "total_employees": total_employees,
        "present_today": present,
        "wfo_today": wfo,
        "wfh_today": wfh,
        "leave_today": approved_leaves,
        "absent_today": absent,
        "late_today": late,
        "not_checked_out": not_checked_out,
        "trend": trend,
        "department_stats": dept_stats,
    }

# ---- Reports ----
async def _report_data(date_from: str, date_to: str, employee_id: Optional[str] = None, department_id: Optional[str] = None):
    q: dict = {"attendance_date": {"$gte": date_from, "$lte": date_to}}
    if employee_id: q["employee_id"] = employee_id
    if department_id:
        emp_ids = [e["id"] for e in await db.employees.find({"department_id": department_id}).to_list(2000)]
        q["employee_id"] = {"$in": emp_ids}
    items = await db.attendance.find(q).sort("attendance_date", 1).to_list(10000)
    emps = {e["id"]: e for e in await db.employees.find().to_list(2000)}
    depts = {d["id"]: d.get("name", "") for d in await db.departments.find().to_list(1000)}
    rows = []
    for x in items:
        c = clean_doc(x)
        e = emps.get(c.get("employee_id")) or {}
        c["employee_name"] = e.get("full_name", "-")
        c["employee_code"] = e.get("employee_id", "-")
        c["department_name"] = depts.get(e.get("department_id"), "-")
        rows.append(c)
    return rows

@api.get("/reports/weekly")
async def weekly_report(week_start: str, employee_id: Optional[str] = None, department_id: Optional[str] = None, user: dict = Depends(get_current_user)):
    start = datetime.fromisoformat(week_start).date()
    end = start + timedelta(days=6)
    if user.get("role") == "employee":
        emp = await db.employees.find_one({"user_id": user["id"]})
        employee_id = emp["id"] if emp else "-"
    rows = await _report_data(start.isoformat(), end.isoformat(), employee_id, department_id)
    return {"period": {"from": start.isoformat(), "to": end.isoformat()}, "rows": rows, "summary": _summarize(rows)}

@api.get("/reports/monthly")
async def monthly_report(year: int, month: int, employee_id: Optional[str] = None, department_id: Optional[str] = None, user: dict = Depends(get_current_user)):
    start = date(year, month, 1)
    if month == 12:
        end = date(year+1, 1, 1) - timedelta(days=1)
    else:
        end = date(year, month+1, 1) - timedelta(days=1)
    if user.get("role") == "employee":
        emp = await db.employees.find_one({"user_id": user["id"]})
        employee_id = emp["id"] if emp else "-"
    rows = await _report_data(start.isoformat(), end.isoformat(), employee_id, department_id)
    return {"period": {"from": start.isoformat(), "to": end.isoformat()}, "rows": rows, "summary": _summarize(rows)}

def _summarize(rows: List[dict]) -> dict:
    return {
        "total": len(rows),
        "wfo": sum(1 for r in rows if r.get("attendance_type") == "WFO"),
        "wfh": sum(1 for r in rows if r.get("attendance_type") == "WFH"),
        "late": sum(1 for r in rows if r.get("attendance_status") == "late"),
        "on_time": sum(1 for r in rows if r.get("attendance_status") == "on_time"),
        "incomplete": sum(1 for r in rows if r.get("check_in_time") and not r.get("check_out_time")),
    }

# ---- Exports ----
def _fmt_time(iso: Optional[str]) -> str:
    if not iso: return "-"
    try:
        return datetime.fromisoformat(iso).strftime("%H:%M")
    except Exception:
        return "-"

@api.get("/exports/excel")
async def export_excel(date_from: str, date_to: str, employee_id: Optional[str] = None, department_id: Optional[str] = None, admin: dict = Depends(require_admin)):
    rows = await _report_data(date_from, date_to, employee_id, department_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Kehadiran"
    headers = ["No", "Tanggal", "NIP", "Nama", "Departemen", "Tipe", "Check-in", "Check-out", "Status", "GPS", "Rencana Kerja", "Output"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = XLFont(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="800000", end_color="800000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, r in enumerate(rows, 1):
        ws.append([
            i, r.get("attendance_date"), r.get("employee_code"), r.get("employee_name"),
            r.get("department_name"), r.get("attendance_type"),
            _fmt_time(r.get("check_in_time")), _fmt_time(r.get("check_out_time")),
            r.get("attendance_status"), "Terverifikasi" if r.get("gps_verified") else "Tidak",
            r.get("activity") or "-", r.get("output") or "-",
        ])
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 12), 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=laporan_kehadiran_{date_from}_{date_to}.xlsx"})

@api.get("/exports/pdf")
async def export_pdf(date_from: str, date_to: str, employee_id: Optional[str] = None, department_id: Optional[str] = None, admin: dict = Depends(require_admin)):
    rows = await _report_data(date_from, date_to, employee_id, department_id)
    settings = await _get_settings_dict()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("t", parent=styles["Title"], textColor=colors.HexColor("#800000"))
    story = [
        Paragraph(settings.get("office_name", "Kantor Pemerintahan"), title_style),
        Paragraph("Laporan Kehadiran Pegawai", styles["Heading2"]),
        Paragraph(f"Periode: {date_from} s/d {date_to}", styles["Normal"]),
        Spacer(1, 12),
    ]
    data = [["No", "Tanggal", "NIP", "Nama", "Departemen", "Tipe", "Masuk", "Pulang", "Status", "GPS"]]
    for i, r in enumerate(rows, 1):
        data.append([
            i, r.get("attendance_date"), r.get("employee_code"), r.get("employee_name"),
            r.get("department_name"), r.get("attendance_type"),
            _fmt_time(r.get("check_in_time")), _fmt_time(r.get("check_out_time")),
            r.get("attendance_status"), "OK" if r.get("gps_verified") else "-",
        ])
    tbl = Table(data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#800000")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("ALIGN", (0,0), (-1,-1), "LEFT"),
        ("FONTSIZE", (0,0), (-1,-1), 8),
        ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.whitesmoke, colors.white]),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 24))
    story.append(Paragraph("Mengetahui,", styles["Normal"]))
    story.append(Spacer(1, 48))
    story.append(Paragraph("(_______________________)", styles["Normal"]))
    doc.build(story)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename=laporan_kehadiran_{date_from}_{date_to}.pdf"})

# ---- Seed ----
async def seed():
    # Settings
    if not await db.settings.find_one({"id": "app"}):
        await db.settings.insert_one(DEFAULT_SETTINGS.copy())

    # Admin
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@gov.id").lower()
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    if not await db.users.find_one({"email": admin_email}):
        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "email": admin_email,
            "password_hash": hash_password(admin_password),
            "role": "admin",
            "status": "active",
            "name": "Administrator",
            "created_at": now_utc_iso(),
        })

    # Departments
    if await db.departments.count_documents({}) == 0:
        depts = [
            {"id": str(uuid.uuid4()), "name": "Sekretariat", "description": "Sekretariat Umum", "status": "active", "created_at": now_utc_iso()},
            {"id": str(uuid.uuid4()), "name": "Kepegawaian", "description": "Bagian Kepegawaian", "status": "active", "created_at": now_utc_iso()},
            {"id": str(uuid.uuid4()), "name": "Keuangan", "description": "Bagian Keuangan", "status": "active", "created_at": now_utc_iso()},
        ]
        await db.departments.insert_many(depts)

    # Location
    if await db.locations.count_documents({}) == 0:
        await db.locations.insert_one({
            "id": str(uuid.uuid4()), "name": "Kantor Pusat", "latitude": -6.175392, "longitude": 106.827153,
            "allowed_radius": 200, "status": "active", "created_at": now_utc_iso(),
        })

    # Employees
    if await db.employees.count_documents({}) == 0:
        depts = await db.departments.find().to_list(10)
        sample = [
            ("EMP001", "Budi Santoso", "budi@gov.id", "Kepala Sub Bagian", depts[0]["id"], "081234567890"),
            ("EMP002", "Siti Nurhaliza", "siti@gov.id", "Staf Kepegawaian", depts[1]["id"], "081234567891"),
            ("EMP003", "Andi Wijaya", "andi@gov.id", "Analis Keuangan", depts[2]["id"], "081234567892"),
        ]
        for eid, name, email, pos, dept, phone in sample:
            uid = str(uuid.uuid4())
            await db.users.insert_one({
                "id": uid, "email": email, "password_hash": hash_password("employee123"),
                "role": "employee", "status": "active", "created_at": now_utc_iso(),
            })
            await db.employees.insert_one({
                "id": str(uuid.uuid4()), "user_id": uid, "employee_id": eid, "full_name": name,
                "email": email, "position": pos, "department_id": dept, "phone": phone,
                "profile_photo": "", "status": "active", "created_at": now_utc_iso(),
            })

    # Holidays
    if await db.holidays.count_documents({}) == 0:
        yr = date.today().year
        for md, nm in [(f"{yr}-01-01", "Tahun Baru"), (f"{yr}-08-17", "Hari Kemerdekaan RI"), (f"{yr}-12-25", "Hari Natal")]:
            await db.holidays.insert_one({"id": str(uuid.uuid4()), "holiday_date": md, "name": nm, "description": "", "created_at": now_utc_iso()})

    # Indexes
    await db.users.create_index("email", unique=True)
    await db.employees.create_index("employee_id", unique=True)
    await db.attendance.create_index([("employee_id", 1), ("attendance_date", 1)])

@api.get("/")
async def root():
    return {"app": "SiKerja API", "status": "ok"}

app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def on_start():
    await seed()

@app.on_event("shutdown")
async def on_shutdown():
    client.close()

logging.basicConfig(level=logging.INFO)
